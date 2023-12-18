import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import os

# Define the headers to read in from various files
fbs_header = ['Facility', 'Booking Date', 'Booking Start Time', 'Booking End Time', 'Booking Owner', 'Purpose']
tms_header = ['Course Title', 'Session Date', 'S-Time', 'E-Time', 'Venue']

# Headers for output
new_tms_header = tms_header.copy()
new_tms_header.append('Remarks')

def read_files():
    """
    Read in files that are starting with `TMS` and `FBS`, and ends with Excel extension
    """
    tms, fbs = {}, {} 
    
    # Read file that starts with `TMS` and `FBS`, and both are Excel files
    for file in os.listdir():
        if (file.startswith("TMS") and file.endswith(".xlsx")):
            tms = pd.read_excel(file, usecols=tms_header).fillna("-").T.to_dict()
        elif (file.startswith("FBS") and file.endswith(".xlsx")):
            fbs = pd.read_excel(file, usecols=fbs_header).fillna("-").T.to_dict()

    # If either of the file is missing, then the code will return "False".
    if not tms or not fbs:
        return False, {}, {}
    else:
        return True, tms, fbs

def change_venue_name(venue):
    """
    Utility function to change the venue to long form for standardisation.
    """

    # Venue mapping to map the different short forms to long forms.
    venue_mapping = {
        "Classroom": "Class Room",
        "SMUC": "SMU Connexion",
        "SMUA Room 1": "Booking needed!",
        "YPHSL": "Yong Pung How School of Law",
        "LKCSB": "Lee Kong Chian School of Business",
        "SOE/SCIS2": "School of Economics/School of Computing & Information Systems 2",
        "SOA": "School of Accountancy",
        "SOSS/CIS": "School of Social Sciences/College of Integrative Studies",
        "SCIS1": "School of Computing & Information System 1",
        "SCIS": "School of Computing & Information System 1"
    }

    # Checks the venue string to see if there are any short forms.
    # If exists, it will be changed.
    for short_form, long_form in venue_mapping.items():
        venue = venue.replace(short_form, long_form)

    return venue


def fbs_tms_title_mapping(fbs, tms):
    """
    Converts some of the data to be of the same type for easier comparison.
    Mainly, this will be changing the venue names from short form to long form.
    """
    course_titles = []

    for key, item in tms.items():
        # Conversion of types of data for standardising and easier comparison
        tms[key]['S-Time'] = dt.datetime.strptime(item['S-Time'], '%I:%M %p').time()
        tms[key]['E-Time'] = dt.datetime.strptime(item['E-Time'], '%I:%M %p').time()
        tms[key]['Session Date'] = pd.to_datetime(item['Session Date']).date()

        # Change venue name from short forms to long forms for standardising
        tms[key]['Venue'] = change_venue_name(item['Venue'])

        if item['Course Title'] not in course_titles:
            course_titles.append(item['Course Title'])
    
    # To store the truncated title as key-value pair so that it is easier to map.
    fbs_titles = {}

    for key, item in fbs.items():
        # Conversion of types of data for standardising and easier comparison
        fbs[key]['Booking Start Time'] = pd.to_datetime(item['Booking Start Time']).time()
        fbs[key]['Booking End Time'] = pd.to_datetime(item['Booking End Time']).time()
        fbs[key]['Booking Date'] = pd.to_datetime(item['Booking Date']).date()

        # Change venue name from short forms to long forms for standardising
        fbs[key]['Facility'] = change_venue_name(item['Facility'])

        purpose = item['Purpose']

        # Checks if the course name has been found before.
        if purpose not in fbs_titles:
            # Goes through all the different records for the `actual` name of the course to match.
            for title in course_titles:
                if purpose in title:
                    fbs_titles[purpose] = title
                    fbs[key]['Purpose'] = title
                    break
        else:
            fbs[key]['Purpose'] = fbs_titles.get(purpose)

    return fbs_titles

def verify_bookings(tms, fbs_dict):
    """
    Verify the booking records in TMS with the FBS booking; TMS records against FBS booking records.

    It tries to check if the same `Course Name`, then `Session Date`.
    If a record could be found, it tries to check if the timing is within the timeframe of the booking.
    """
    res = []

    # Loops through each TMS record to verify if booking is found or booking in the TMS record is the same as FBS record
    for value in tms.values():
        # Gets the `Course Title`
        key = value.get('Course Title')

        # Preparing the data for output based on TMS values
        val = [key, value.get('Session Date'), value.get('S-Time').strftime('%I:%M %p'), value.get('E-Time').strftime('%I:%M %p'), value.get('Venue')]

        # Cannot find the course name in FBS / the name is mismatched
        if key not in fbs_dict:
            val.append('Not found in FBS List / Name mismatched')
        else:
            fbs_value = fbs_dict[key]
            
            # Formats the `Session Date` to be string value.
            session_date = str(value['Session Date'])

            # Cannot find the date of this booking
            if session_date not in fbs_value:
                val.append('Booking is missing for this record')
            else:
                # Search through the starting time as the start time is not exactly the same for each course.
                for fbs_start_time, fbs_item in fbs_value[session_date].items():

                    # Check start time earlier or equal to the course start time, and
                    # check end time is later or equal to the course end time
                    if fbs_start_time <= str(value['S-Time']) and fbs_item['End'] >= str(value['E-Time']):
                        # Check TMS venue name against FBS booking venue for the course and session date
                        if fbs_item['Venue'] == value['Venue']:
                            val.append('Venue matched')
                        # If the venue is Online class
                        elif value['Venue'] == 'Online Class':
                            val.append('No booking needed')
                        # If the venue is mismatched
                        else:
                            val.append('Venue NOT matched')
                    # If the start or end time is beyond the booking time
                    else:
                        val.append('Timing exceeds booking')

        res.append(val)

    return res

if __name__ == "__main__":
    valid, tms, fbs = read_files()
    if not valid:
        exit("Files are missing")

    # Formats the title and other relevant fields for comparison
    fbs_titles = fbs_tms_title_mapping(fbs, tms)

    fbs_dict = {}

    # Convert to dictionary for easier finding and structure of data
    for item in fbs.values():
        title_item = fbs_dict.get(item['Purpose'], {})
        title_item_start_date = title_item.get(item['Booking Date'], {})

        title_item_start_date[str(item['Booking Start Time'])] = {'Venue': item['Facility'], 'End': str(item['Booking End Time'])}
        title_item[str(item['Booking Date'])] = title_item_start_date
        fbs_dict[item['Purpose']] = title_item

    # Sort the data to be ascending order, according to the course name
    fbs_dict = dict(sorted(fbs_dict.items()))

    res = verify_bookings(tms, fbs_dict)

    filename = 'output.xlsx'

    # -------------- FORMATTING & OUTPUTTING OF DATA ----------------
    data_df = pd.DataFrame(res, columns=new_tms_header)

    writer = pd.ExcelWriter(filename, engine="xlsxwriter")
    data_df.to_excel(writer, sheet_name='Sheet1', index=False)

    worksheet = writer.sheets['Sheet1']

    header_format = writer.book.add_format({'bold': True, 'fg_color': "#808080", 'border': 1, 'font_size': 15})
    normal_text = writer.book.add_format({'text_wrap': True})

    # To color the header column and bold it
    for colno, value in enumerate(data_df.columns.values):
        worksheet.write(0, colno, value, header_format)

    worksheet.set_column('A:A', 40, normal_text)
    worksheet.set_column('B:B', 20, normal_text)
    worksheet.set_column('C:D', 10, normal_text)
    worksheet.set_column('E:F', 30, normal_text)

    writer.close()

    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']

    booking_needed_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    venue_mismatch_fill = PatternFill(start_color='FE8780', end_color='FE8780', fill_type='solid')
    time_exceed_fill = PatternFill(start_color='BA92BE', end_color='BA92BE', fill_type='solid')
    course_not_found_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    

    for row in ws.iter_rows(min_row=2, values_only=False):
        remarks = row[5].value

        if remarks == 'Venue NOT matched':
            for cell in row:
                cell.fill = venue_mismatch_fill
        elif remarks == 'Timing exceeds booking':
            for cell in row:
                cell.fill = time_exceed_fill
        elif remarks == 'Booking is missing for this record':
            for cell in row:
                cell.fill = booking_needed_fill
        elif remarks == 'Not found in FBS List / Name mismatched':
            for cell in row:
                cell.fill = course_not_found_fill 

    wb.save(filename)
