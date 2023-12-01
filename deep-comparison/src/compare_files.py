import copy
from datetime import datetime as dt
from deepdiff import DeepDiff as dd
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import os

# headers for reading in values from the original CDL files
headers = ['Pillar', 'Course No.', 'Course Title', 'Status', 'Course Run ID',\
       'Mode of Delivery', 'Type of Runs (Public or Corporate)', 'Start Date',\
       'End Date', 'Session Date & Time', 'Session Venue', 'Location by Date',\
       'Total no. of sessions', 'Registered Pax', 'Enrolled Pax', 'Total Pax', 'Venue Category']

new_cols = copy.deepcopy(headers)
new_cols.extend(['Last Updated', "Changes From", "Changes To"])

# get current directory
path = os.getcwd()
# get parent directory
parent_dir = os.path.abspath(os.path.join(path, os.pardir)) + "/"

# get files with file name that starts with `CDL` and ends with `.xlsx` (Excel extension)
# sort them in order
files = sorted([parent_dir + f for f in os.listdir(parent_dir) if f.startswith('CDL') and f.endswith('.xlsx')])
files_df = []   # Storing DataFrame of files
files_dict = [] # Storing Dictionary of files
datetime_now = dt.now().strftime("%Y-%m-%d %H:%M")
new_row = {}


def read_files():
    """
        Read files as Dataframe and convert into Dictionary.
        Stores the result in an array of Dictionary.
    """
    for i in files:
        file_read = pd.read_excel(i, usecols=headers, converters={"Total Pax": int})
        files_df.append(file_read)

    files_dict.append(files_df[0].set_index("Course No.").T.to_dict())
    files_dict.append(files_df[1].set_index("Course No.").T.to_dict())

def update_files_last_update(datetime_now):
    """
        Update CDL files with last update time
    """
    for i in range(len(files_df)):
        files_df[i].loc[:, ['Last Updated']] = f'Last Updated: {datetime_now}'

        writer = pd.ExcelWriter(files[i], engine="xlsxwriter")
        files_df[i].to_excel(writer, sheet_name='Sheet1', index=False)

        worksheet = writer.sheets["Sheet1"]

        normal_text = writer.book.add_format({'text_wrap': True})
        bold_text = writer.book.add_format({'text_wrap': True, 'bold': True, 'font_size': 15})
        header_format = writer.book.add_format({'bold': True, 'fg_color': "#ffcccc", 'border': 1, 'font_size': 15})


        # To color the header column and bold it
        for colno, value in enumerate(files_df[i].columns.values):
            worksheet.write(0, colno, value, header_format)

        worksheet.set_column('A:A', 20, normal_text)
        worksheet.set_column('B:C', 40, normal_text)
        worksheet.set_column('D:G', 20, normal_text)
        worksheet.set_column('H:I', 20, bold_text)
        worksheet.set_column('J:L', 40, bold_text)
        worksheet.set_column('M:M', 20, normal_text)
        worksheet.set_column('N:P', 20, bold_text)
        worksheet.set_column('Q:R', 20, normal_text)

        writer.close()


def check_differences():
    """
        Check the differences between the new and the old files.
        Record down the modified rows, and note down the newly added rows.
    """
    modified_row = {}

    for k, v in files_dict[1].items():
        # new value
        if k not in files_dict[0]:
            new_row[k] = v

    diff = dd(files_dict[0], files_dict[1], ignore_order=True)

    for key, value in diff['values_changed'].items():
        course_no = key.split("']['")[0].split("['")[1]
        column_key = key.split("']['")[1].split("']")[0]
        temp = modified_row.get(course_no, [])
        if temp == []:
            temp.append(f'{column_key} changed from \n{value["old_value"]}')
            temp.append(f'{column_key} changed to \n{value["new_value"]}')
        else:
            temp[0] += f'\n\n{column_key} changed from \n{value["old_value"]}'
            temp[1] += f'\n\n{column_key} changed to \n{value["new_value"]}'

        modified_row[course_no] = temp
    
    return modified_row


def structure_data(modified_row={}):
    """
        Structure the data according to the format of the original files
    """
    res = []
    for k, v in files_dict[1].items():
        temp = []
        for header in headers:
            if header == 'Course No.':
                temp.append(k)
            else:
                temp.append(v[header])
        if k in new_row:
            temp.extend([f'Last updated: {datetime_now}', "New Row", "New Row"])
        elif k not in modified_row:
            temp.extend([f'Last updated: {datetime_now}', "Not modified", "Not modified"])
        else:
            temp.extend([f'Last updated: {datetime_now}', modified_row[k][0], modified_row[k][1]])
        res.append(temp)

    return res


def export_to_file():
    """
        Export the new data into the file with formatting
    """
    df = pd.DataFrame(new_data, columns=new_cols)
    df = df.sort_values(['Start Date', 'Course No.'])
    file_names = []

    for file in files:
        file_names.append(file.split("CDL_")[1].split(".")[0])
    
    export_filename = "Combined_CDL_" + file_names[0] + "-" + file_names[1] + ".xlsx"

    writer = pd.ExcelWriter(parent_dir + export_filename, engine="xlsxwriter")
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    worksheet = writer.sheets["Sheet1"]

    normal_text = writer.book.add_format({'text_wrap': True})
    bold_text = writer.book.add_format({'text_wrap': True, 'bold': True, 'font_size': 15})
    header_format = writer.book.add_format({'bold': True, 'fg_color': "#ffcccc", 'border': 1, 'font_size': 15})


    # To color the header column and bold it
    for colno, value in enumerate(df.columns.values):
        worksheet.write(0, colno, value, header_format)

    worksheet.set_column('A:A', 20, normal_text)
    worksheet.set_column('B:C', 40, normal_text)
    worksheet.set_column('D:G', 20, normal_text)
    worksheet.set_column('H:I', 20, bold_text)
    worksheet.set_column('J:L', 40, bold_text)
    worksheet.set_column('K:K', 80, bold_text)
    worksheet.set_column('M:M', 20, normal_text)
    worksheet.set_column('N:P', 20, bold_text)
    worksheet.set_column('Q:Q', 20, normal_text)
    worksheet.set_column('R:R', 20, bold_text)
    worksheet.set_column('S:T', 60, bold_text)

    writer.close()

    wb = openpyxl.load_workbook(filename=parent_dir+export_filename)
    ws = wb['Sheet1']
    conditional_col = 1
    fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type="solid")
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_value = row[conditional_col].value

        # Check if condition is met
        if cell_value in new_row:
            for cell in row:
                cell.fill = fill
    wb.save(parent_dir + export_filename)


"""
    Starting point of the Python code
"""
if __name__ == '__main__':
    if not len(files) == 2:
        exit("There must be exactly 2 Excel files.")
    
    read_files()

    if files_dict[0] != files_dict[1]:
        modified_row = check_differences()
        new_data = structure_data(modified_row)
        export_to_file()
    else:
        new_data = files_df[1]
        new_data['Changes From'] = 'Not modified'
        new_data['Changes To'] = 'Not modified'
        new_data['Last Updated'] = f'Last updated: {datetime_now}'
        export_to_file()
        
    update_files_last_update(datetime_now)
